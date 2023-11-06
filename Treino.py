# from openpyxl import Workbook
#
# wb = Workbook()
#
# nome_arquivo = 'primeiro_arquivo.xlsx'
#
# ws1 = wb.active
# ws1.title = 'Planilha 1'
#
# dados = [
#      ['Ano', 'Lucro', 'Custos'],
#     [2021, '30%', '40%'],
#     [2021, '30%', '40%'],
#     [2021, '30%', '40%']
# ]
# for linha in dados:
#     ws1.append(linha)
#
# ws1['H13']= 13
# wb.save(filename=nome_arquivo)

from openpyxl import load_workbook

arquivo = 'C:\\Users\\silva.valdenir\\Downloads\\nomes.xlsx'
wb = load_workbook(arquivo)

planilha = wb['Planilha1']

print(planilha['A2'].value)